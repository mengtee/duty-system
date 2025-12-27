"""
Export Service - Export schedules to Excel format.
Demonstrates File Generation and Data Processing capabilities.
"""
from typing import List, Optional
from datetime import date
from sqlalchemy.orm import Session
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from models.schedule import Schedule
from models.schedule_assignment import ScheduleAssignment
import os
from datetime import datetime

class ExportService:
    """
    Service class for exporting schedules to various formats.
    Currently supports Excel (XLSX) export.
    """
    
    def __init__(self, db: Session):
        """
        Initialize ExportService with database session.
        
        Args:
            db: Database session
        """
        self.db = db
    
    def export_to_excel(self, start_date: Optional[date] = None, end_date: Optional[date] = None) -> str:
        """
        Export schedules to Excel file.
        
        Args:
            start_date: Start date filter
            end_date: End date filter
            
        Returns:
            Path to the generated Excel file
        """
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "排班表"
        
        # Set column widths
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 30
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Add title
        ws.merge_cells('A1:E1')
        title_cell = ws['A1']
        title_cell.value = "值班排班表"
        title_cell.font = Font(bold=True, size=16)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30
        
        # Add export info
        ws.merge_cells('A2:E2')
        info_cell = ws['A2']
        date_range = ""
        if start_date and end_date:
            date_range = f"({start_date.isoformat()} 至 {end_date.isoformat()})"
        elif start_date:
            date_range = f"(从 {start_date.isoformat()})"
        elif end_date:
            date_range = f"(至 {end_date.isoformat()})"
        info_cell.value = f"导出时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} {date_range}"
        info_cell.alignment = Alignment(horizontal="center")
        
        # Add headers
        headers = ['日期', '班次', '值班人员', '状态', '备注']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Query schedules
        query = self.db.query(Schedule)
        if start_date:
            query = query.filter(Schedule.schedule_date >= start_date)
        if end_date:
            query = query.filter(Schedule.schedule_date <= end_date)
        
        schedules = query.order_by(Schedule.schedule_date).all()
        
        # Add data
        row = 4
        for schedule in schedules:
            assignments = schedule.get_assignments()
            
            if assignments:
                for assignment in assignments:
                    ws.cell(row=row, column=1).value = schedule.schedule_date.isoformat()
                    ws.cell(row=row, column=2).value = self._translate_shift_type(schedule.shift_type)
                    ws.cell(row=row, column=3).value = assignment.staff.name if assignment.staff else ""
                    ws.cell(row=row, column=4).value = self._translate_status(assignment.status)
                    ws.cell(row=row, column=5).value = assignment.notes or ""
                    
                    # Apply borders
                    for col in range(1, 6):
                        ws.cell(row=row, column=col).border = border
                        ws.cell(row=row, column=col).alignment = Alignment(horizontal="center", vertical="center")
                    
                    row += 1
            else:
                # Empty schedule
                ws.cell(row=row, column=1).value = schedule.schedule_date.isoformat()
                ws.cell(row=row, column=2).value = self._translate_shift_type(schedule.shift_type)
                ws.cell(row=row, column=3).value = "未分配"
                ws.cell(row=row, column=4).value = "-"
                ws.cell(row=row, column=5).value = ""
                
                for col in range(1, 6):
                    ws.cell(row=row, column=col).border = border
                    ws.cell(row=row, column=col).alignment = Alignment(horizontal="center", vertical="center")
                
                row += 1
        
        # Save file
        filename = f"schedule_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(os.getcwd(), filename)
        wb.save(filepath)
        
        return filepath
    
    def _translate_shift_type(self, shift_type: str) -> str:
        """Translate shift type to Chinese."""
        translations = {
            'morning': '早班',
            'afternoon': '中班',
            'night': '晚班',
            '全天': '全天'
        }
        return translations.get(shift_type, shift_type)
    
    def _translate_status(self, status: str) -> str:
        """Translate status to Chinese."""
        translations = {
            'scheduled': '已排班',
            'completed': '已完成',
            'cancelled': '已取消'
        }
        return translations.get(status, status)
